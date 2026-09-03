"""
fatture/fiscale.py — Blocco C: gestione fiscale (dashboard, spese P.IVA,
parametri, export Excel).

Rotte HTML:
  GET /fatture/situazione            -> dashboard fiscale (riepilogo + scadenze)
  GET /fatture/spese-piva            -> lista movimenti P.IVA
  GET /fatture/spese-piva/nuova      -> form nuovo movimento
  GET /fatture/spese-piva/<int:id>   -> form edit movimento
  GET /fatture/parametri             -> editor parametri fiscali

Rotte JSON:
  GET    /fatture/api/situazione?anno=YYYY
  GET    /fatture/api/export/xlsx?anno=YYYY
  GET    /fatture/api/parametri
  PATCH  /fatture/api/parametri
  GET    /fatture/api/spese-piva
  POST   /fatture/api/spese-piva
  PATCH  /fatture/api/spese-piva/<int:id>
  DELETE /fatture/api/spese-piva/<int:id>

Logica di calcolo verificata sul file di riferimento fornito dall'utente
(Situazione_forfait_Bertoli_Andrea.xlsx, fogli "2026" e "Template"):
  - l'acconto all'80% ("metodo storico") si applica SOLO all'INPS gestione
    separata. L'acconto dell'imposta sostitutiva e' pari al 100% del saldo
    (nessuna riduzione).
  - Scadenza 30/06/anno+1: saldo imposta + saldo INPS + commercialista
    dell'anno + bollo dell'anno.
  - Scadenza 30/11/anno+1: acconto imposta (=saldo, 100%) + acconto INPS
    (=saldo INPS * 80%).
"""
import io
from datetime import date

from flask import Response, request, jsonify, send_file

from . import fatture_bp
from . import accantonamento as acc
from .costanti import (CATEGORIE_SPESE_PIVA, MESI_NOMI, STATI_EMESSE,
                       TIPI_SPESE_PIVA)
from shared.theme import render_page
from shared.design import icon as _icon, info as _info
from shared.supabase_client import get_client, is_configured
from shared.fmt import eur as _fmt_eur, data_it as _fmt_date, pct


PARAMETRI_DEFAULT = {
    "id": 1, "regime": "RF19", "ateco": "622010",
    "ateco_descrizione": "Attività di consulenza informatica",
    "coeff_ateco": 0.67, "aliquota_imposta": 0.05, "aliquota_inps": 0.2607,
    "aliquota_acconto": 0.80, "bollo_soglia": 77.47, "bollo_importo": 2.00,
    "limite_fatturato_anno": 85000, "data_apertura_piva": "2026-05-28",
    "anno_fine_regime_agevolato": 2031,
    # Quanto vale una giornata da 8 ore: e' quello che la precompilazione
    # dal timesheet moltiplica per le giornate del mese (README §8.14).
    "tariffa_giornaliera": 250.00,
    **acc.PARAMETRI_DEFAULT,
}

PARAMETRI_CAMPI = (
    "regime", "coeff_ateco", "aliquota_imposta", "aliquota_inps",
    "aliquota_acconto", "bollo_soglia", "bollo_importo",
    "limite_fatturato_anno", "data_apertura_piva", "anno_fine_regime_agevolato",
    "tariffa_giornaliera",
) + acc.PARAMETRI_CAMPI

# Paracadute contro l'errore di battitura, non un vincolo di legge:
# coeff_ateco/aliquote sono percentuali 0-1 (non 0-100), margine_sicurezza
# ha un tetto largo (200%) solo per intercettare un "50" digitato per
# sbaglio al posto di "0.50". Senza questo controllo un valore fuori scala
# produce un accantonamento assurdo (negativo, oltre il lordo) visibile
# ovunque nell'app — l'unico punto che clampa un movimento di denaro reale
# e' giroconto.calcola(), ma solo li'.
PARAMETRI_LIMITI = {
    "coeff_ateco":                (0.0, 1.0),
    "aliquota_imposta":           (0.0, 1.0),
    "aliquota_inps":              (0.0, 1.0),
    "aliquota_acconto":           (0.0, 1.0),
    "bollo_soglia":               (0.0, None),
    "bollo_importo":              (0.0, None),
    "limite_fatturato_anno":      (0.0, None),
    "anno_fine_regime_agevolato": (2000, 2100),
    "tariffa_giornaliera":        (0.0, None),
    "margine_sicurezza":          (0.0, 2.0),
    "costi_fissi_annui":          (0.0, None),
    "fatturato_atteso_anno":      (0.0, None),
    "acconto_imposta_perc":       (0.0, 1.0),
    "acconto_prima_rata_perc":    (0.0, 1.0),
}


def _valida_parametri(payload: dict) -> str | None:
    """Messaggio d'errore se un valore e' fuori dai limiti di sanita', None
    se tutto passa."""
    for campo, (minimo, massimo) in PARAMETRI_LIMITI.items():
        if campo not in payload or payload[campo] is None:
            continue
        try:
            v = float(payload[campo])
        except (TypeError, ValueError):
            return f'"{campo}" deve essere un numero.'
        if minimo is not None and v < minimo:
            return f'"{campo}" non può essere minore di {minimo}.'
        if massimo is not None and v > massimo:
            return (f'"{campo}" = {v} è fuori dal range plausibile (max {massimo}) '
                    f'— controlla se intendevi una percentuale come 0-1 invece di 0-100.')
    return None


MOVIMENTO_CAMPI = (
    "data", "importo", "tipo", "descrizione", "categoria",
    "sottocategoria", "fattura_id", "ricorrente", "note",
)


def _supabase_or_error():
    if not is_configured():
        return None, ('<div class="notice warn">Supabase non configurato.</div>')
    return get_client(), None


def _esc(v) -> str:
    return (str(v) if v is not None else "").replace("&", "&amp;").replace(
        "<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


# ---------------------------------------------------------------------------
# Calcolo situazione fiscale
# ---------------------------------------------------------------------------

# Aliquota standard dell'imposta sostitutiva dopo il quinquennio agevolato
# (art. 1 c. 64 L. 190/2014). L'agevolata (5 %, default in PARAMETRI_DEFAULT)
# vale solo per i primi 5 anni di attivita'.
ALIQUOTA_IMPOSTA_STANDARD = 0.15


def _aliquota_imposta_per_anno(param: dict, anno: int) -> float:
    """
    param["anno_fine_regime_agevolato"] segna il primo anno in cui l'aliquota
    torna quella standard. Senza questo controllo il campo "Primo anno con
    aliquota al 15%" nel form Parametri resterebbe solo testo: nessun
    calcolo lo leggerebbe mai, e l'app continuerebbe a usare il 5% anche
    dopo la fine dell'agevolazione finche' qualcuno non se ne accorge e
    non lo cambia a mano.
    """
    try:
        fine_agevolato = int(param.get("anno_fine_regime_agevolato"))
    except (TypeError, ValueError):
        fine_agevolato = None
    if fine_agevolato is not None and anno >= fine_agevolato:
        return ALIQUOTA_IMPOSTA_STANDARD
    return float(param.get("aliquota_imposta") or 0.05)


def _get_parametri(sb) -> dict:
    try:
        r = sb.table("b2f_parametri_fiscali").select("*").eq("id", 1).single().execute()
        param = {**PARAMETRI_DEFAULT, **r.data} if r.data else dict(PARAMETRI_DEFAULT)
    except Exception:
        param = dict(PARAMETRI_DEFAULT)
    # Corretta per l'anno corrente: chi chiama senza un anno esplicito
    # (giroconto su un incasso di oggi, anteprima nell'editor, form
    # Parametri) intende sempre "quanto si applica adesso".
    param["aliquota_imposta"] = _aliquota_imposta_per_anno(param, date.today().year)
    return param


def _mesi_attivita(anno: int, data_apertura_iso: str) -> int:
    """Mesi ragguagliati di attivita' nell'anno indicato.

    Regola: se il giorno di apertura e' successivo al 15, il mese di
    apertura non conta ai fini del ragguaglio (coerente con la regola
    Agenzia Entrate per il calcolo del limite forfettario).
    """
    try:
        y, m, d = (int(x) for x in data_apertura_iso[:10].split("-"))
    except Exception:
        return 12
    if anno < y:
        return 0
    if anno > y:
        return 12
    mese_apertura = m + 1 if d > 15 else m
    return max(12 - mese_apertura + 1, 0)


def _situazione_data(sb, anno: int) -> dict:
    param = _get_parametri(sb)
    coeff = float(param["coeff_ateco"])
    # Non il valore che _get_parametri() ha gia' corretto per l'anno di
    # oggi: qui "anno" puo' essere un anno diverso da quello corrente (si
    # puo' guardare la situazione di un anno passato o futuro dal
    # selettore), quindi l'aliquota va ricalcolata proprio su quell'anno.
    aliq_imp = _aliquota_imposta_per_anno(param, anno)
    # Riscritta anche dentro param: acc.scomponi() piu' sotto e l'export
    # Excel leggono param["aliquota_imposta"], non la variabile locale
    # aliq_imp. Senza questa riga userebbero ancora la correzione fatta
    # da _get_parametri() per l'anno di oggi, non per "anno".
    param["aliquota_imposta"] = aliq_imp
    aliq_inps = float(param["aliquota_inps"])
    aliq_acc = float(param["aliquota_acconto"])
    limite_anno = float(param["limite_fatturato_anno"])

    fatturato_mese = {m: 0.0 for m in range(1, 13)}
    bollo_mese = {m: 0.0 for m in range(1, 13)}
    rivalsa_mese = {m: 0.0 for m in range(1, 13)}
    incasso_mese = {m: 0.0 for m in range(1, 13)}
    commercialista_mese = {m: 0.0 for m in range(1, 13)}
    spese_piva_uscite_tot = 0.0

    try:
        r = (sb.table("b2f_fatture")
               .select("data,data_incasso,totale,bollo,bollo_addebitato,cassa_importo,stato")
               .gte("data", f"{anno}-01-01").lte("data", f"{anno}-12-31")
               .in_("stato", list(STATI_EMESSE)).execute())
        for f in (r.data or []):
            mese = int(f["data"][5:7])
            fatturato_mese[mese] += float(f.get("totale") or 0)
            if f.get("bollo_addebitato"):
                bollo_mese[mese] += float(f.get("bollo") or 0)
            # Solo per trasparenza nell'export: la rivalsa resta dentro
            # "fatturato"/"totale" (concorre al reddito, vedi accantonamento.py),
            # qui si somma a parte solo per mostrarne la quota.
            rivalsa_mese[mese] += float(f.get("cassa_importo") or 0)
    except Exception:
        pass

    try:
        # Incassato del mese: filtrato sulla DATA di incasso, non sullo
        # stato. Da quando l'incasso e' un passo in mezzo al percorso
        # (costanti.py), una fattura pagata prosegue verso lo studio e lo
        # SDI: chiedere stato='incassata' farebbe sparire dai calcoli per
        # cassa proprio le fatture piu' avanti nel giro. Il filtro di data
        # esclude gia' da solo chi non ha incassato.
        r = (sb.table("b2f_fatture").select("data_incasso,totale")
               .neq("stato", "annullata")
               .gte("data_incasso", f"{anno}-01-01").lte("data_incasso", f"{anno}-12-31")
               .execute())
        for f in (r.data or []):
            if not f.get("data_incasso"):
                continue
            mese = int(f["data_incasso"][5:7])
            incasso_mese[mese] += float(f.get("totale") or 0)
    except Exception:
        pass

    try:
        r = (sb.table("b2f_spese_piva").select("data,importo,categoria,tipo")
               .eq("tipo", "uscita")
               .gte("data", f"{anno}-01-01").lte("data", f"{anno}-12-31")
               .execute())
        for s in (r.data or []):
            spese_piva_uscite_tot += float(s.get("importo") or 0)
            if s.get("categoria") == "commercialista":
                mese = int(s["data"][5:7])
                commercialista_mese[mese] += float(s.get("importo") or 0)
    except Exception:
        pass

    mensile = []
    tot = {k: 0.0 for k in ("fatturato", "imponibile", "incasso", "imposta",
                             "inps_saldo", "inps_acconto", "bollo", "commercialista",
                             "rivalsa")}
    for m in range(1, 13):
        fatt = fatturato_mese[m]
        imponibile = round(fatt * coeff, 2)
        inps_saldo = round(imponibile * aliq_inps, 2)
        # L'imposta sostitutiva si calcola sul reddito al netto dei contributi
        # INPS deducibili: (Imponibile - INPS Saldo) * aliquota.
        imposta = round((imponibile - inps_saldo) * aliq_imp, 2)
        inps_acconto = round(inps_saldo * aliq_acc, 2)
        incasso = incasso_mese[m]
        bollo = bollo_mese[m]
        comm = commercialista_mese[m]
        rivalsa = rivalsa_mese[m]
        # "Stipendio" del foglio di riferimento: incasso meno tutto, acconti
        # inclusi. Resta per parita' con l'export Excel.
        netto = round(incasso - imposta - inps_saldo - inps_acconto - bollo - comm, 2)
        # Quanto resta davvero di competenza: gli acconti sono un anticipo
        # sull'anno successivo, non un costo dell'anno.
        netto_comp = round(incasso - imposta - inps_saldo - bollo - comm, 2)
        mensile.append({
            "mese": m, "nome": MESI_NOMI[m - 1],
            "fatturato": fatt, "imponibile": imponibile, "incasso": incasso,
            "imposta": imposta, "inps_saldo": inps_saldo, "inps_acconto": inps_acconto,
            "bollo": bollo, "commercialista": comm, "rivalsa": rivalsa,
            "netto": netto, "netto_competenza": netto_comp,
        })
        tot["fatturato"] += fatt
        tot["imponibile"] += imponibile
        tot["incasso"] += incasso
        tot["imposta"] += imposta
        tot["inps_saldo"] += inps_saldo
        tot["rivalsa"] += rivalsa
        tot["inps_acconto"] += inps_acconto
        tot["bollo"] += bollo
        tot["commercialista"] += comm

    for k in tot:
        tot[k] = round(tot[k], 2)

    mesi_attivita = _mesi_attivita(anno, param["data_apertura_piva"])
    limite_ragguagliato = round(limite_anno / 12 * mesi_attivita, 2)

    # L'acconto dell'imposta sostitutiva e' pari al 100 % del saldo (nessuna
    # riduzione), a differenza dell'INPS che va all'80 % col metodo storico.
    acc_imposta_perc = float(param.get("acconto_imposta_perc") or 1.0)
    imposta_acconto = round(tot["imposta"] * acc_imposta_perc, 2)

    # L'acconto si versa in DUE rate, non tutto a novembre: 40 % con il
    # saldo di giugno (di norma prorogato al 31 luglio) e il resto il 30
    # novembre. La quota e' un parametro (README §8.17): a 0 si torna al
    # modello di prima. La differenza non e' contabile ma di cassa, ed e'
    # esattamente il numero che serve sapere — quanto avere liquido a
    # giugno.
    prima_rata_perc = min(max(float(param.get("acconto_prima_rata_perc") or 0), 0.0), 1.0)
    acconti_tot = round(imposta_acconto + tot["inps_acconto"], 2)
    acconto_prima_rata = round(acconti_tot * prima_rata_perc, 2)
    acconto_seconda_rata = round(acconti_tot - acconto_prima_rata, 2)

    scadenza_giugno = round(tot["imposta"] + tot["inps_saldo"]
                            + tot["commercialista"] + tot["bollo"]
                            + acconto_prima_rata, 2)
    scadenza_novembre = acconto_seconda_rata

    # Due grandezze che prima erano confuse in un unico "netto stimato":
    #  - competenza: quanto dell'anno resta davvero tuo. Gli acconti non
    #    sono un costo, sono un anticipo che si scomputa dal saldo dopo.
    #  - cassa: quanto serve avere da parte per onorare le due scadenze,
    #    che nell'anno di transizione cadono sullo stesso anno solare.
    netto_competenza = round(tot["incasso"] - tot["imposta"] - tot["inps_saldo"]
                             - tot["bollo"] - tot["commercialista"], 2)
    dovuto_saldo = round(tot["imposta"] + tot["inps_saldo"], 2)
    acconti = round(imposta_acconto + tot["inps_acconto"], 2)
    cassa_da_riservare = round(scadenza_giugno + scadenza_novembre, 2)

    # Accantonamento consigliato sull'incassato dell'anno, esposto anche
    # dall'API cosi' le altre pagine non devono rifare il calcolo.
    scomposizione = acc.scomponi(tot["incasso"], param,
                                 fatturato_riferimento=tot["incasso"],
                                 rivalsa=tot["rivalsa"], bollo_addebitato=tot["bollo"],
                                 anno=anno)

    return {
        "anno": anno,
        "parametri": param,
        "accantonamento": {
            **scomposizione["importi"],
            "aliquote": scomposizione["aliquote"],
            "netti": scomposizione["netti"],
        },
        "limite_ragguagliato": limite_ragguagliato,
        "mesi_attivita": mesi_attivita,
        "totali": {
            "fatturato": tot["fatturato"],
            "imponibile": tot["imponibile"],
            "incasso": tot["incasso"],
            "imposta_accantonata": tot["imposta"],
            "inps_saldo_accantonato": tot["inps_saldo"],
            "inps_acconto_accantonato": tot["inps_acconto"],
            "imposta_acconto": imposta_acconto,
            "bollo_totale": tot["bollo"],
            "rivalsa_totale": tot["rivalsa"],
            "commercialista_totale": tot["commercialista"],
            "spese_piva_totali": round(spese_piva_uscite_tot, 2),
            # Nuove grandezze, esplicite
            "dovuto_saldo": dovuto_saldo,
            "acconti": acconti,
            "acconto_prima_rata": acconto_prima_rata,
            "acconto_seconda_rata": acconto_seconda_rata,
            "scadenza_giugno": scadenza_giugno,
            "scadenza_novembre": scadenza_novembre,
            "cassa_da_riservare": cassa_da_riservare,
            "netto_competenza": netto_competenza,
            # Alias storici, mantenuti per non rompere chi li leggeva
            "totale_da_versare": cassa_da_riservare,
            "netto_stimato": netto_competenza,
        },
        "mensile": mensile,
        "scadenze": [
            {"data": f"{anno + 1}-06-30",
             "descrizione": ("Saldo imposta + saldo INPS + commercialista + bollo"
                             + (f" + 1ª rata acconti ({int(prima_rata_perc * 100)} %)"
                                if acconto_prima_rata > 0 else "")),
             "importo": scadenza_giugno,
             "voci": [
                 ("Saldo imposta sostitutiva", tot["imposta"]),
                 ("Saldo INPS gestione separata", tot["inps_saldo"]),
                 ("Commercialista dell'anno", tot["commercialista"]),
                 ("Bollo dell'anno", tot["bollo"]),
                 (f"1ª rata acconti {anno + 1}", acconto_prima_rata),
             ]},
            {"data": f"{anno + 1}-11-30",
             "descrizione": f"2ª rata acconti ({int((1 - prima_rata_perc) * 100)} %)",
             "importo": scadenza_novembre,
             "voci": [
                 (f"2ª rata acconti {anno + 1}", acconto_seconda_rata),
             ]},
        ],
    }


def saldo_piva(sb, al: str | None = None) -> dict:
    """
    Saldo reale del conto P.IVA a una data (default: oggi).

    Non e' il saldo dell'anno filtrato che mostra /fatture/spese-piva: e'
    quanto c'e' sul conto, tutti i movimenti dall'apertura fino ad `al`.
    Dal secondo anno in poi i due numeri divergono, perche' il conto non
    riparte da zero a gennaio.

    Le tre direzioni sono diverse fra loro:
      entrata    il cliente ha pagato          -> entra
      uscita     un costo della P.IVA          -> esce
      giroconto  la tua quota va sul personale -> esce (ma non e' un costo)

    `rivalsa_incassata` e' la quota di rivalsa INPS gia' entrata su questo
    conto insieme agli incassi: non e' un ricavo tuo, e' la parte del
    corrispettivo destinata al contributo previdenziale. Va lasciata qui:
    serve a farla vedere, perche' dentro il saldo e' invisibile.

    Le righe si leggono a blocchi: PostgREST tronca ogni richiesta a un
    tetto, e un saldo troncato sarebbe sbagliato senza dare errore.
    """
    al = al or date.today().isoformat()
    vuoto = {"al": al, "disponibile": False, "saldo": 0.0, "entrate": 0.0,
             "uscite": 0.0, "girati": 0.0, "rivalsa_incassata": 0.0,
             "movimenti": 0}

    entrate = uscite = girati = 0.0
    n = 0
    offset, passo = 0, 1000
    while True:
        try:
            r = (sb.table("b2f_spese_piva").select("importo,tipo,data")
                   .lte("data", al).order("data", desc=False)
                   .range(offset, offset + passo - 1).execute())
            pagina = r.data or []
        except Exception:
            return vuoto
        for m in pagina:
            imp = abs(float(m.get("importo") or 0))
            tipo = m.get("tipo")
            if tipo == "entrata":
                entrate += imp
            elif tipo == "uscita":
                uscite += imp
            elif tipo == "giroconto":
                girati += imp
            n += 1
        if len(pagina) < passo:
            break
        offset += passo

    # Rivalsa gia' incassata: si conta sulle fatture incassate, non sui
    # movimenti P.IVA, perche' e' la fattura a sapere quanta parte del
    # totale era rivalsa (`cassa_importo`).
    rivalsa = 0.0
    try:
        r = (sb.table("b2f_fatture").select("cassa_importo,data_incasso")
               .neq("stato", "annullata").lte("data_incasso", al).execute())
        rivalsa = sum(float(f.get("cassa_importo") or 0) for f in (r.data or [])
                      if f.get("data_incasso"))
    except Exception:
        rivalsa = 0.0

    return {
        "al": al,
        "disponibile": True,
        "entrate": round(entrate, 2),
        "uscite": round(uscite, 2),
        "girati": round(girati, 2),
        "rivalsa_incassata": round(rivalsa, 2),
        "saldo": round(entrate - uscite - girati, 2),
        "movimenti": n,
    }


# Alias pubblici: app.py e i moduli fratelli hanno bisogno di questi due.
def get_parametri(sb) -> dict:
    """Parametri fiscali correnti, con i default applicati."""
    return _get_parametri(sb)


def situazione_data(sb, anno: int) -> dict:
    """Situazione fiscale completa dell'anno indicato."""
    return _situazione_data(sb, anno)


# ---------------------------------------------------------------------------
# Dashboard /fatture/situazione
# ---------------------------------------------------------------------------

def _fondo_tasse(sb, anno: int, fabbisogno: float) -> dict:
    """
    Il fondo tasse dell'anno: quanto servira', quanto c'e' davvero.

    E' la domanda che nessuna pagina faceva. La card della fattura dice
    quanto accantonare *su quella fattura*; nessuno sommava le decisioni
    e le confrontava con i soldi veri. Dieci fatture ripartite ciascuna
    in modo difendibile possono lasciare un buco che si scopre a giugno,
    perche' il conto P.IVA si svuota un giroconto alla volta e ogni
    singolo giroconto sembra piccolo.

    Tre numeri, e il terzo e' l'unico che non si puo' raccontare:
      serve       il fabbisogno sull'incassato dell'anno (tasse + costi)
      deciso      la somma degli accantonamenti sulle fatture ripartite
      sul conto   il saldo P.IVA reale, oggi

    "deciso" e' un'intenzione, "sul conto" e' un fatto: se divergono, il
    denaro e' uscito da qualche altra parte.
    """
    dati = {
        "serve": round(float(fabbisogno or 0), 2),
        "deciso": 0.0, "da_ripartire": 0.0,
        "n_ripartite": 0, "n_da_ripartire": 0,
        "saldo_piva": None, "disponibile": False,
    }
    try:
        r = (sb.table("b2f_fatture")
               .select("totale,accantonamento_importo,data_giroconto,data_incasso,stato")
               .neq("stato", "annullata")
               .gte("data_incasso", f"{anno}-01-01")
               .lte("data_incasso", f"{anno}-12-31").execute())
        righe = r.data or []
    except Exception:
        return dati

    for f in righe:
        if not f.get("data_incasso"):
            continue
        if f.get("data_giroconto"):
            dati["deciso"] += float(f.get("accantonamento_importo") or 0)
            dati["n_ripartite"] += 1
        else:
            # Non ripartita: l'incasso e' ancora tutto sul conto P.IVA.
            # Non e' "accantonato" — e' solo denaro non ancora diviso.
            dati["da_ripartire"] += float(f.get("totale") or 0)
            dati["n_da_ripartire"] += 1

    dati["deciso"] = round(dati["deciso"], 2)
    dati["da_ripartire"] = round(dati["da_ripartire"], 2)
    try:
        dati["saldo_piva"] = float(saldo_piva(sb).get("saldo") or 0)
        dati["disponibile"] = True
    except Exception:
        pass
    return dati


def _scadenze_tutte(sb, param: dict, s_anno: dict, anno: int) -> list:
    """
    Tutte le scadenze conosciute, in ordine di data.

    Ogni anno con incassi ne genera due: il 30 giugno dell'anno dopo
    (saldo + prima rata degli acconti, piu' commercialista e bollo di
    quell'anno) e il 30 novembre (seconda rata). Gli anni futuri non
    compaiono: senza incassi non c'e' niente da calcolare, e una stima
    inventata qui sarebbe peggio di una riga in meno.

    Per l'anno in vista si usano gli importi gia' calcolati da
    `_situazione_data` — sono gli stessi numeri delle altre card, e due
    conti diversi sulla stessa scadenza sarebbero un bug in attesa.
    """
    from . import accantonamento as acc

    scadenze = []
    for sc in (s_anno.get("scadenze") or []):
        if sc["importo"] > 0:
            scadenze.append({**sc, "anno_comp": anno})

    # Gli altri anni: incassato per anno, dalle fatture non annullate.
    try:
        r = (sb.table("b2f_fatture")
               .select("totale,bollo,bollo_addebitato,data_incasso,stato")
               .neq("stato", "annullata").execute())
        righe = r.data or []
    except Exception:
        righe = []

    per_anno: dict[int, dict] = {}
    for f in righe:
        di = f.get("data_incasso")
        if not di:
            continue
        try:
            y = int(str(di)[:4])
        except (TypeError, ValueError):
            continue
        if y == anno:
            continue          # gia' coperto dagli importi precisi sopra
        v = per_anno.setdefault(y, {"incasso": 0.0, "bollo": 0.0})
        v["incasso"] += float(f.get("totale") or 0)
        if f.get("bollo_addebitato"):
            v["bollo"] += float(f.get("bollo") or 0)

    for y, v in sorted(per_anno.items()):
        if v["incasso"] <= 0:
            continue
        sc = acc.scomponi(v["incasso"], param, fatturato_riferimento=v["incasso"],
                          anno=y)
        giugno = round(sc["entro_giugno"] + v["bollo"], 2)
        if giugno > 0:
            scadenze.append({
                "data": f"{y + 1}-06-30", "anno_comp": y,
                "descrizione": f"Saldo {y} + 1ª rata acconti {y + 1}",
                "importo": giugno,
                "voci": [(f"Saldo {y} (INPS + imposta)", sc["saldo"]),
                         (f"1ª rata acconti {y + 1}", sc["acconto_prima_rata"]),
                         ("Bollo dell'anno", v["bollo"])],
            })
        if sc["entro_novembre"] > 0:
            scadenze.append({
                "data": f"{y + 1}-11-30", "anno_comp": y,
                "descrizione": f"2ª rata acconti {y + 1}",
                "importo": sc["entro_novembre"],
                "voci": [(f"2ª rata acconti {y + 1}", sc["entro_novembre"])],
            })

    scadenze.sort(key=lambda x: x["data"])
    return scadenze


def _cascata(scadenze: list, disponibile: float | None) -> list:
    """
    Versa il denaro disponibile nelle scadenze, in ordine di data.

    La prima si riempie fino all'orlo, poi comincia la seconda, e cosi'
    via: e' come si comportano i soldi veri, che non sanno di essere
    destinati a una scadenza piuttosto che a un'altra. Il denaro e' uno
    solo — il saldo del conto P.IVA, cuscinetto compreso — e la domanda a
    cui questo risponde e' "fin dove arrivo".
    """
    resto = None if disponibile is None else max(float(disponibile), 0.0)
    out = []
    for sc in scadenze:
        importo = float(sc["importo"] or 0)
        if resto is None:
            coperto = None
        else:
            coperto = round(min(resto, importo), 2)
            resto = round(resto - coperto, 2)
        manca = None if coperto is None else round(max(importo - coperto, 0.0), 2)
        quota = (coperto / importo) if (coperto is not None and importo > 0) else None
        out.append({**sc, "coperto": coperto, "manca": manca, "quota": quota})
    return out, (resto if resto is not None else None)


_MESI_SCAD = ("gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
              "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre")


def _card_calendario(sb, s: dict, anno: int, saldo_oggi: float | None) -> str:
    """
    Un blocco per scadenza: quanto si paga, quanto ne ho da parte, quanto
    manca. Il denaro disponibile riempie i blocchi in ordine di data.

    Le altre card rispondono a "quanto"; questa risponde a "quando, e ci
    arrivo?". Il fabbisogno non serve tutto insieme e non e' destinato in
    partenza a una scadenza piuttosto che a un'altra: c'e' un mucchio di
    soldi sul conto P.IVA, e la domanda vera e' fin dove arriva. Per
    questo il riempimento e' a cascata — la prima scadenza si copre per
    intero, poi comincia la seconda — e non una divisione proporzionale,
    che darebbe quattro blocchi tutti mezzi pieni e nessuna informazione.

    Visibili di default le scadenze non ancora passate; le altre stanno
    dietro il bottone, che appare solo se c'e' davvero altro da mostrare.
    """
    scadenze = _scadenze_tutte(sb, param=s["parametri"], s_anno=s, anno=anno)
    if not scadenze:
        return ""

    scadenze, avanzo = _cascata(scadenze, saldo_oggi)
    oggi = date.today().isoformat()
    totale = round(sum(x["importo"] for x in scadenze), 2)
    premessa = _info(
        "Il denaro sul conto P.IVA non &egrave; destinato in partenza a una "
        "scadenza piuttosto che a un&apos;altra: &egrave; un mucchio solo, "
        "cuscinetto compreso. Qui lo si versa nei blocchi <strong>in ordine "
        "di data</strong> — la prima si copre per intero, poi comincia la "
        "seconda. La percentuale dice fin dove arrivi.")

    # Quelle passate restano consultabili ma non in mezzo ai piedi.
    nascoste = [i for i, x in enumerate(scadenze) if x["data"] < oggi]

    blocchi = []
    for i, sc in enumerate(scadenze):
        y, m, d = sc["data"].split("-")
        passata = sc["data"] < oggi
        quota = sc["quota"]
        if quota is None:
            cls, pct_txt, barra, riga_stato = "", "", "", ""
        else:
            cls = "pos" if quota >= 0.999 else ("warn" if quota >= 0.5 else "neg")
            pct_txt = f"{quota * 100:.0f}".replace(".", ",")
            barra = (f'<div class="meter"><i class="{cls}" '
                     f'style="width:{min(quota * 100, 100):.1f}%"></i></div>')
            if sc["manca"] <= 0:
                riga_stato = (f'<div class="scb-stato pos">Coperta &mdash; '
                              f'&euro; {_fmt_eur(sc["coperto"])} da parte</div>')
            else:
                manca_pct = f"{(1 - quota) * 100:.0f}".replace(".", ",")
                riga_stato = (
                    f'<div class="scb-stato {cls}">Manca il <strong>{manca_pct} %</strong>'
                    f' &mdash; &euro; {_fmt_eur(sc["manca"])} da trovare</div>')

        voci = [(n, v) for n, v in (sc.get("voci") or []) if v]
        voci_html = "".join(
            f'<div class="scb-voce"><span>{n}</span>'
            f'<span class="tnum">&euro; {_fmt_eur(v)}</span></div>'
            for n, v in voci)

        blocchi.append(f'''
        <div class="scb{" passata" if passata else ""}"
             data-scb="{i}"{' hidden' if i in nascoste else ''}>
          <div class="scb-head">
            <div class="scb-data">
              <span class="scb-giorno">{int(d)}</span>
              <span class="scb-mese">{_MESI_SCAD[int(m) - 1]}</span>
              <span class="scb-anno">{y}</span>
            </div>
            <div class="scb-cifre">
              <div class="scb-tot tnum">&euro; {_fmt_eur(sc["importo"])}</div>
              <div class="scb-desc">{sc["descrizione"]}</div>
            </div>
            <div class="scb-pct tnum {cls}">{pct_txt + " %" if pct_txt else ""}</div>
          </div>
          {barra}
          {riga_stato}
          <div class="scb-voci">{voci_html}</div>
        </div>''')

    bottone = ""
    if nascoste:
        bottone = (
            f'<button type="button" class="btn ghost block mt-3" id="scbPiu">'
            f'{_icon("calendar")}Mostra anche le {len(nascoste)} scadenze gi&agrave; '
            f'passate</button>'
            '<script>(function(){'
            'var b=document.getElementById("scbPiu");if(!b)return;'
            'b.addEventListener("click",function(){'
            'document.querySelectorAll("[data-scb][hidden]").forEach(function(e){'
            'e.hidden=false});b.remove();});})();</script>')

    if saldo_oggi is None:
        coda = ""
    elif avanzo and avanzo > 0:
        coda = (f'<div class="small mt-3"><strong class="pos">Coperte tutte, '
                f'e avanzano &euro; {_fmt_eur(avanzo)}.</strong> Quella parte '
                f'&egrave; davvero tua.</div>')
    else:
        scoperte = [x for x in scadenze if x["manca"] and x["manca"] > 0]
        mancante = round(sum(x["manca"] for x in scoperte), 2)
        coda = (f'<div class="small mt-3"><strong class="neg">Mancano '
                f'&euro; {_fmt_eur(mancante)}</strong> per coprire '
                f'{"l&apos;ultima scadenza" if len(scoperte) == 1 else f"le ultime {len(scoperte)} scadenze"}.</div>')

    return f'''
    <style>
      .scb{{padding:var(--sp-3);border-radius:var(--r-sm);
        background:var(--surface-3);margin-bottom:var(--sp-3)}}
      .scb:last-of-type{{margin-bottom:0}}
      .scb.passata{{opacity:.62}}
      .scb-head{{display:flex;gap:var(--sp-3);align-items:flex-start}}
      .scb-data{{flex:none;width:58px;text-align:center;
        display:flex;flex-direction:column;line-height:1.15}}
      .scb-giorno{{font-family:var(--display);font-size:23px;color:var(--ink)}}
      .scb-mese{{font-size:11.5px;color:var(--ink-2)}}
      .scb-anno{{font-size:11px;color:var(--ink-3)}}
      .scb-cifre{{flex:1;min-width:0}}
      .scb-tot{{font-size:18px;font-weight:600;color:var(--ink);line-height:1.15}}
      .scb-desc{{font-size:11.5px;color:var(--ink-3);margin-top:1px}}
      .scb-pct{{flex:none;font-size:15px;font-weight:600}}
      .scb-pct.pos{{color:var(--pos)}}
      .scb-pct.warn{{color:var(--warn)}}
      .scb-pct.neg{{color:var(--neg)}}
      .scb .meter{{margin-top:var(--sp-3)}}
      .scb-stato{{font-size:12px;margin-top:6px}}
      .scb-stato.pos{{color:var(--pos)}}
      .scb-stato.warn{{color:var(--warn)}}
      .scb-stato.neg{{color:var(--neg)}}
      .scb-voci{{margin-top:var(--sp-2);display:flex;flex-direction:column;gap:3px}}
      .scb-voce{{display:flex;justify-content:space-between;gap:10px;
        font-size:12px;color:var(--ink-2)}}
    </style>
    <div class="card">
      <div class="card-head">
        <div class="eyebrow">Le scadenze, e quanto ho da parte{premessa}</div>
        <span class="chip">&euro; {_fmt_eur(totale)} in tutto</span>
      </div>
      {"".join(blocchi)}
      {bottone}
      {coda}
    </div>'''


def _card_fondo_tasse(d: dict, anno: int) -> str:
    """
    La card del fondo. Il numero grande e' la copertura, non un importo:
    la domanda e' "ci sono?", e la risposta e' una percentuale.
    """
    if not d.get("disponibile") or d["serve"] <= 0:
        return ""

    saldo = d["saldo_piva"] or 0.0
    serve = d["serve"]
    copertura = saldo / serve if serve > 0 else 0.0
    manca = round(max(serve - saldo, 0.0), 2)
    avanza = round(max(saldo - serve, 0.0), 2)

    if copertura >= 1.0:
        cls, chip = "pos", "Coperto"
    elif copertura >= 0.7:
        cls, chip = "warn", "Quasi"
    else:
        cls, chip = "neg", "Scoperto"

    pct_barra = min(copertura * 100, 100)
    pct_txt = f"{copertura * 100:.0f}".replace(".", ",")

    if manca > 0:
        verdetto = (f'<strong class="neg">Mancano &euro; {_fmt_eur(manca)}.</strong> '
                    f'Sono soldi che a giugno serviranno comunque: se non sono '
                    f'sul conto P.IVA, arriveranno da quello personale.')
    else:
        verdetto = (f'<strong class="pos">Ci sono tutti, e avanzano '
                    f'&euro; {_fmt_eur(avanza)}.</strong> Quella parte &egrave; '
                    f'davvero tua: &egrave; il cuscinetto che hai accumulato '
                    f'scegliendo gli scenari.')

    righe_dettaglio = f'''
      <div class="row"><span class="t">Ti serviranno{_info(
          f"Saldo {anno}, acconti {anno + 1} e costi fissi, calcolati "
          f"sull&apos;incassato dell&apos;anno.")}</span>
        <span class="v tnum">&euro; {_fmt_eur(serve)}</span></div>
      <div class="row"><span class="t">Sul conto P.IVA, oggi{_info(
          "Entrate meno uscite meno giroconti, dall&apos;apertura della "
          "partita IVA. &Egrave; il saldo vero del conto, non quello "
          "dell&apos;anno filtrato.")}</span>
        <span class="v tnum {cls}">&euro; {_fmt_eur(saldo)}</span></div>
      <div class="row"><span class="t">Di cui deciso di tenere{_info(
          f"La quota rimasta sul conto dalle {d['n_ripartite']} fatture "
          f"gi&agrave; ripartite. &Egrave; un&apos;intenzione: quello che c&apos;&egrave; "
          f"davvero &egrave; la riga sopra.")}</span>
        <span class="v tnum">&euro; {_fmt_eur(d["deciso"])}</span></div>
      <div class="row"><span class="t">Incassato non ancora ripartito{_info(
          f"{d['n_da_ripartire']} fatture incassate senza giroconto: i soldi "
          f"sono tutti sul conto P.IVA, la quota tua non &egrave; ancora uscita. "
          f"Quando la sposterai, il saldo qui sopra scender&agrave;.")}</span>
        <span class="v tnum">&euro; {_fmt_eur(d["da_ripartire"])}</span></div>'''

    return f'''
    <div class="card">
      <div class="card-head">
        <div class="eyebrow">Fondo tasse {anno}{_info(
          "Il conto P.IVA porta anche il fondo degli anni prima. Se il saldo "
          "dell&apos;anno scorso non l&apos;hai ancora versato, una parte di quel "
          "denaro &egrave; gi&agrave; impegnata e la copertura qui accanto &egrave; "
          "pi&ugrave; ottimista del vero.")}</div>
        <span class="chip {cls}">{chip}</span>
      </div>
      <div class="stat">
        <div class="val tnum {cls}">{pct_txt} %</div>
        <div class="lbl">di quello che ti servir&agrave; &egrave; sul conto P.IVA</div>
      </div>
      <div class="meter mt-3"><i class="{cls}" style="width:{pct_barra:.1f}%"></i></div>
      <div class="small mt-3">{verdetto}</div>
      <div class="rows detail mt-3">{righe_dettaglio}</div>

    </div>'''



@fatture_bp.get("/situazione")
def situazione_dashboard():
    sb, err = _supabase_or_error()
    breadcrumb = [("Fatture", "/fatture"), ("Situazione fiscale", "")]
    if err:
        return _render(err, breadcrumb=breadcrumb)

    anno_default = date.today().year
    anno = request.args.get("anno", type=int) or anno_default

    try:
        s = _situazione_data(sb, anno)
    except Exception as e:
        return _render(f'<div class="notice err">Errore: {str(e)[:200]}</div>',
                       breadcrumb=breadcrumb)

    t = s["totali"]
    # Il limite dei 85.000 € e' per cassa (art. 1 c. 54 L. 190/2014, "ricavi
    # o compensi percepiti"): a differenza di imposta/INPS/scadenze qui sotto
    # (tenute apposta sul fatturato emesso, per restare allineate al foglio
    # Excel di riferimento — vedi il box sotto), questo indicatore specifico
    # deve seguire l'incassato, o il rischio "stai per uscire dal regime"
    # risulterebbe falsato proprio a cavallo di fine anno, quando conta di piu'.
    pct_limite = 0.0
    if s["limite_ragguagliato"] > 0:
        pct_limite = min(t["incasso"] / s["limite_ragguagliato"] * 100, 100)
    limite_cls = "neg" if pct_limite >= 90 else "warn" if pct_limite >= 70 else ""
    pct_it = f"{pct_limite:.1f}".replace(".", ",")

    anno_opts = "".join(
        f'<option value="{y}"{" selected" if y == anno else ""}>{y}</option>'
        for y in range(anno_default + 1, anno_default - 5, -1)
    )
    toolbar_anno = (
        '<div class="toolbar">'
        '<select class="select-pill" aria-label="Anno"'
        ' onchange="location.href=\'/fatture/situazione?anno=\'+this.value">'
        f'{anno_opts}</select></div>'
    )

    # --- Tessere principali -------------------------------------------------
    kpi = f'''
    <div class="grid kpi mb-4">
      <div class="card"><div class="stat">
        <div class="val tnum">€ {_fmt_eur(t["fatturato"], 0)}</div>
        <div class="lbl">Fatturato {anno}</div></div></div>
      <div class="card"><div class="stat">
        <div class="val tnum">€ {_fmt_eur(t["incasso"], 0)}</div>
        <div class="lbl">Incassato</div>
        <div class="hint">su cui si pagano le tasse</div></div></div>
      <div class="card"><div class="stat">
        <div class="val tnum accent">€ {_fmt_eur(t["cassa_da_riservare"], 0)}</div>
        <div class="lbl">Cassa da riservare</div>
        <div class="hint">saldo + acconti</div></div></div>
      <div class="card"><div class="stat">
        <div class="val tnum pos">€ {_fmt_eur(t["netto_competenza"], 0)}</div>
        <div class="lbl">Netto di competenza</div>
        <div class="hint">quanto ti resta davvero</div></div></div>
    </div>

    <details class="explain card mb-4">
      <summary>Due basi di calcolo diverse, ed è voluto</summary>
      <p class="small muted mt-2">
        Imposta, INPS e scadenze qui sopra sono calcolate sul <strong>fatturato
        emesso</strong>, per restare allineate al foglio Excel di riferimento.
        L'accantonamento qui sotto lavora invece sull'<strong>incassato</strong>,
        che è la base corretta del forfettario (regime di cassa). Finché emetti
        e incassi nello stesso anno i due numeri quasi coincidono; a cavallo di
        fine anno divergono, ed è normale che sia così.
      </p>
    </details>'''

    # --- Accantonamento sull'incassato dell'anno ----------------------------
    scomposizione = acc.scomponi(t["incasso"], s["parametri"],
                                 fatturato_riferimento=t["incasso"],
                                 rivalsa=t.get("rivalsa_totale", 0),
                                 bollo_addebitato=t.get("bollo_totale", 0),
                                 anno=anno)
    acc_card = acc.card_html(
        scomposizione,
        titolo=f"Da accantonare sul {anno}",
        contesto="Calcolato sull'incassato dell'anno. Il minimo è il dovuto "
                 "esatto; il sicuro copre anche l'anno in cui saldo e acconti "
                 "cadono insieme.",
        uid="accAnno",
        anno_saldo=anno, anno_acconto=anno + 1,
    ) if t["incasso"] > 0 else ""

    # --- Il fondo tasse: quanto servira', quanto c'e' davvero ---------------
    # La card sopra dice quanto accantonare. Questa dice se i soldi ci
    # sono. Sono due domande diverse, e finora si rispondeva solo alla
    # prima — una per fattura, senza che nessuno le sommasse mai.
    fondo = _fondo_tasse(sb, anno, scomposizione.get("fabbisogno_con_costi", 0))
    fondo_card = _card_fondo_tasse(fondo, anno) if t["incasso"] > 0 else ""

    # --- Il calendario: le due scadenze, e cosa resta dopo ciascuna ---------
    calendario_card = _card_calendario(
        sb, s, anno,
        fondo.get("saldo_piva") if fondo.get("disponibile") else None)

    # --- Limite forfettario --------------------------------------------------
    residuo = max(s["limite_ragguagliato"] - t["incasso"], 0)
    limite = f'''
    <div class="card">
      <div class="card-head">
        <div class="eyebrow">Limite forfettario</div>
        <span class="chip {limite_cls or "accent"}">{pct_it} %</span>
      </div>
      <div class="meter"><i class="{limite_cls}" style="width:{pct_limite:.1f}%"></i></div>
      <div class="small muted mt-3">
        € {_fmt_eur(t["incasso"])} incassati su € {_fmt_eur(s["limite_ragguagliato"])}
        ragguagliati a {s["mesi_attivita"]} mesi di attività — per cassa, come
        vuole la norma sul limite.
        Restano <strong>€ {_fmt_eur(residuo)}</strong>.
      </div>
    </div>'''

    # --- Riepilogo mensile: solo i mesi con movimento -------------------------
    mesi_attivi = [m for m in s["mensile"]
                   if m["fatturato"] or m["incasso"] or m["commercialista"]]
    if mesi_attivi:
        righe = "".join(f'''
        <tr>
          <td>{m["nome"]}</td>
          <td class="num">{_fmt_eur(m["fatturato"])}</td>
          <td class="num">{_fmt_eur(m["incasso"])}</td>
          <td class="num">{_fmt_eur(m["inps_saldo"])}</td>
          <td class="num">{_fmt_eur(m["imposta"])}</td>
          <td class="num">{_fmt_eur(m["netto_competenza"])}</td>
        </tr>''' for m in mesi_attivi)
        tabella = f'''
        <div class="scroll-x">
          <table class="table">
            <thead><tr>
              <th>Mese</th><th class="num">Fatturato</th><th class="num">Incassato</th>
              <th class="num">INPS</th><th class="num">Imposta</th>
              <th class="num">Netto</th>
            </tr></thead>
            <tbody>{righe}</tbody>
            <tfoot><tr>
              <td>Totale</td>
              <td class="num">{_fmt_eur(t["fatturato"])}</td>
              <td class="num">{_fmt_eur(t["incasso"])}</td>
              <td class="num">{_fmt_eur(t["inps_saldo_accantonato"])}</td>
              <td class="num">{_fmt_eur(t["imposta_accantonata"])}</td>
              <td class="num">{_fmt_eur(t["netto_competenza"])}</td>
            </tr></tfoot>
          </table>
        </div>'''
    else:
        tabella = ('<div class="small muted">Nessun movimento registrato '
                   f'per il {anno}.</div>')

    mensile = f'''
    <div class="card">
      <div class="card-head"><div class="eyebrow">Riepilogo mensile</div></div>
      {tabella}
    </div>'''

    azioni = f'''
    <div class="card">
      <div class="card-head"><div class="eyebrow">Strumenti</div></div>
      <div class="actions col mt-0">
        <a class="btn" href="/fatture/api/export/xlsx?anno={anno}">
          {_icon("download")}Esporta Excel
        </a>
        <a class="btn ghost" href="/fatture/parametri">Parametri fiscali</a>
        <a class="btn ghost" href="/fatture/spese-piva">Movimenti P.IVA</a>
      </div>
    </div>'''

    body = f'''
    {toolbar_anno}
    {kpi}
    {acc_card}
    {fondo_card}
    {calendario_card}
    <div class="grid split mt-4">
      <div class="stack">{mensile}</div>
      <div class="stack">{limite}{azioni}</div>
    </div>
    '''

    return _render(body, eyebrow="Fiscale",
                   title_html='Situazione <em>fiscale</em>',
                   breadcrumb=breadcrumb)


@fatture_bp.get("/api/situazione")
def api_situazione():
    sb, err = _supabase_or_error()
    if err: return jsonify({"error": "supabase not configured"}), 503
    anno = request.args.get("anno", type=int) or date.today().year
    try:
        return jsonify(_situazione_data(sb, anno))
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

def _build_workbook(sb, anno: int):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    s = _situazione_data(sb, anno)
    param = s["parametri"]

    try:
        r = sb.table("b2f_emittente").select("*").eq("id", 1).single().execute()
        em = r.data or {}
    except Exception:
        em = {}
    nome = f'{em.get("nome","") or ""} {em.get("cognome","") or ""}'.strip() or "—"

    FONT = "Georgia"
    money_fmt = '#,##0.00" €"'
    pct_fmt = "0.00%"
    thin = Side(style="thin")
    input_fill = PatternFill("solid", fgColor="FFFFF2CC")

    wb = Workbook()
    ws = wb.active
    ws.title = str(anno)
    ws.sheet_view.showGridLines = False

    def cell(coord, value, bold=False, fmt=None, fill=None, border=None, size=11):
        c = ws[coord]
        c.value = value
        c.font = Font(name=FONT, bold=bold, size=size)
        if fmt: c.number_format = fmt
        if fill: c.fill = fill
        if border: c.border = border
        return c

    # --- Header (righe 1-5) ---
    cell("A1", nome, bold=True)
    cell("A2", param.get("ateco_descrizione") or "")
    cell("A3", f'CODICE ATECO {param.get("ateco","")}')

    cell("E1", "Commercialista", bold=True)
    cell("F1", anno, bold=True)
    cell("G1", round(s["totali"]["commercialista_totale"], 2), fmt=money_fmt)
    cell("H1", "Regime", bold=True)
    cell("I1", float(param["limite_fatturato_anno"]), fmt=money_fmt)
    cell("J1", 12)

    cell("E2", "Bollo Fatture >", bold=True)
    cell("F2", float(param["bollo_soglia"]), fmt=money_fmt)
    cell("G2", float(param["bollo_importo"]), fmt=money_fmt)
    cell("H2", "Attività", bold=True)
    cell("I2", "=I1/J1*J2", fmt=money_fmt)
    cell("J2", s["mesi_attivita"])

    cell("H3", "Residuo limite", bold=True)
    cell("I3", "=I2-SUM(B7:B18)", fmt=money_fmt)

    cell("D4", "COEFFICIENTE", bold=True)
    cell("E4", "Aliquota", bold=True)
    cell("F4", "INPS", bold=True)
    cell("G4", "Acconto", bold=True)

    ws.merge_cells("A5:C5")
    aliq_imp_pct = int(round(float(param["aliquota_imposta"]) * 100))
    cell("A5", f"ALIQUOTA {aliq_imp_pct}% PER I PRIMI 5 ANNI", bold=True)
    cell("D5", float(param["coeff_ateco"]), fmt="0%")
    cell("E5", float(param["aliquota_imposta"]), fmt="0%")
    cell("F5", float(param["aliquota_inps"]), fmt=pct_fmt)
    cell("G5", float(param["aliquota_acconto"]), fmt="0%")

    # --- Riga 6: intestazioni tabella ---
    # "Rivalsa INPS" e' in coda e non fra Fatturato e Imponibile apposta:
    # e' gia' dentro "Fatturato" (concorre al reddito, vedi accantonamento.py
    # e Risposta Agenzia Entrate 428/2022) — qui e' solo la quota mostrata
    # a parte per trasparenza, non un valore da sommare a se'.
    headers = ["Mese", "Fatturato", "Imponibile", "Incasso", "Imposta",
               "INPS Saldo", "INPS Acconto", "Bollo Fattura", "Commercialista",
               "Stipendio", "di cui Rivalsa INPS"]
    header_border = Border(bottom=thin)
    for i, h in enumerate(headers):
        col = chr(ord("A") + i)
        cell(f"{col}6", h, bold=True, border=header_border)

    # --- Righe 7-18: 12 mesi ---
    for i, m in enumerate(s["mensile"]):
        r = 7 + i
        cell(f"A{r}", m["nome"])
        cell(f"B{r}", m["fatturato"], fmt=money_fmt, fill=input_fill)
        cell(f"C{r}", f"=B{r}*$D$5", fmt=money_fmt)
        cell(f"D{r}", m["incasso"], fmt=money_fmt, fill=input_fill)
        cell(f"E{r}", f"=(C{r}-F{r})*$E$5", fmt=money_fmt)
        cell(f"F{r}", f"=C{r}*$F$5", fmt=money_fmt)
        cell(f"G{r}", f"=F{r}*$G$5", fmt=money_fmt)
        cell(f"H{r}", m["bollo"], fmt=money_fmt, fill=input_fill)
        cell(f"I{r}", m["commercialista"], fmt=money_fmt, fill=input_fill)
        cell(f"J{r}", f"=D{r}-E{r}-F{r}-G{r}-H{r}-I{r}", fmt=money_fmt)
        cell(f"K{r}", m["rivalsa"], fmt=money_fmt, fill=input_fill)

    # --- Riga 19: Totale ---
    top_border = Border(top=thin, bottom=Side(style="double"))
    for col in "BCDEFGHIJK":
        cell(f"{col}19", f"=SUM({col}7:{col}18)", bold=True, fmt=money_fmt, border=top_border)
    cell("A19", "Totale", bold=True, border=top_border)

    # --- Righe 21-26: Scadenze ---
    box = Border(top=thin, bottom=thin, left=thin, right=thin)
    cell("A21", "Scadenze", bold=True)
    cell("B21", f"Giugno {anno + 1}", bold=True)
    cell("C21", f"Novembre {anno + 1}", bold=True)
    cell("D21", "Totale", bold=True)

    cell("A22", "Imposta")
    cell("B22", "=E19", fmt=money_fmt)
    cell("C22", "=E19", fmt=money_fmt)

    cell("A23", "INPS")
    cell("B23", "=F19", fmt=money_fmt)
    cell("C23", "=G19", fmt=money_fmt)

    cell("A24", "Commercialista")
    cell("B24", "=I19", fmt=money_fmt)

    cell("A25", "Bollo fatture")
    cell("B25", "=H19", fmt=money_fmt)

    cell("A26", "Totale", bold=True)
    cell("B26", "=SUM(B22:B25)", bold=True, fmt=money_fmt)
    cell("C26", "=SUM(C22:C25)", bold=True, fmt=money_fmt)
    cell("D26", "=B26+C26", bold=True, fmt=money_fmt)

    for row in ws["A21:D26"]:
        for c in row:
            c.border = box

    # Larghezza colonne
    widths = {"A": 22, "B": 13, "C": 13, "D": 13, "E": 12, "F": 12,
              "G": 13, "H": 14, "I": 15, "J": 13, "K": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@fatture_bp.get("/api/export/xlsx")
def api_export_xlsx():
    sb, err = _supabase_or_error()
    if err: return jsonify({"error": "supabase not configured"}), 503
    anno = request.args.get("anno", type=int) or date.today().year
    try:
        buf = _build_workbook(sb, anno)
    except Exception as e:
        return jsonify({"error": str(e)[:250]}), 500
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Situazione_forfait_{anno}.xlsx",
    )


# ---------------------------------------------------------------------------
# Parametri fiscali
# ---------------------------------------------------------------------------

@fatture_bp.get("/parametri")
def parametri_editor():
    sb, err = _supabase_or_error()
    breadcrumb = [("Fatture", "/fatture"), ("Situazione fiscale", "/fatture/situazione"),
                  ("Parametri", "")]
    if err:
        return _render(err, breadcrumb=breadcrumb)
    p = _get_parametri(sb)

    # Anteprima dal vivo: mostra subito l'effetto dei parametri su 1.000 €
    anteprima = acc.scomponi(1000.0, p, fatturato_riferimento=0.0,
                             anno=date.today().year)
    # Le righe dell'anteprima si generano da SCENARI: aggiungerne uno non
    # deve lasciare qui tre nomi scritti a mano che invecchiano in
    # silenzio (e' quello che era successo con "minimo" e "sicuro").
    righe_anteprima = "".join(
        f'''<div class="row"><span class="t">{acc.ETICHETTE[k][0]}{_info(acc.ETICHETTE[k][1])}</span>
           <span class="v tnum{" accent" if k == p.get("scenario_preferito") else ""}">
             € {_fmt_eur(anteprima["importi"][k])}</span></div>'''
        for k in acc.SCENARI)
    scen_opts = "".join(
        f'<option value="{k}"{" selected" if p.get("scenario_preferito") == k else ""}>'
        f'{acc.ETICHETTE[k][0]}</option>' for k in acc.SCENARI
    )

    body = f'''
    <div class="grid split">
      <div class="stack">

        <div class="card">
          <div class="card-head"><div class="eyebrow">Regime</div></div>
          <div class="field"><label>Regime fiscale</label>
            <select id="f_regime">
              <option value="RF19"{" selected" if p["regime"] == "RF19" else ""}>RF19 — Forfettario</option>
            </select></div>
          <div class="field"><label>Codice ATECO{_info("Modificabile solo da database.")}</label>
            <input value="{p['ateco']} — {p.get('ateco_descrizione') or ''}" disabled></div>
          <div class="field-group">
            <div class="field"><label>Data apertura P.IVA</label>
              <input type="date" id="f_data_apertura" value="{p['data_apertura_piva']}"></div>
            <div class="field"><label>Fine regime agevolato{_info("Primo anno con aliquota al 15 %.")}</label>
              <input type="number" id="f_anno_fine" value="{p.get('anno_fine_regime_agevolato') or ''}"></div>
          </div>
        </div>

        <div class="card">
          <div class="card-head"><div class="eyebrow">Aliquote</div></div>
          <div class="field-group">
            <div class="field"><label>Coefficiente redditività{_info("0,67 per la consulenza informatica.")}</label>
              <input type="number" step="0.0001" id="f_coeff" value="{p['coeff_ateco']}"></div>
            <div class="field"><label>Imposta sostitutiva{_info("0,05 nei primi cinque anni.")}</label>
              <input type="number" step="0.0001" id="f_aliq_imp" value="{p['aliquota_imposta']}"></div>
          </div>
          <div class="field-group">
            <div class="field"><label>INPS gestione separata</label>
              <input type="number" step="0.0001" id="f_aliq_inps" value="{p['aliquota_inps']}"></div>
            <div class="field"><label>Acconto INPS{_info("0,80 col metodo storico.")}</label>
              <input type="number" step="0.0001" id="f_aliq_acc" value="{p['aliquota_acconto']}"></div>
          </div>
          <div class="field-group">
            <div class="field"><label>Acconto imposta{_info("1,00 = 100 % del saldo, nessuna riduzione.")}</label>
              <input type="number" step="0.01" id="f_acc_imp" value="{p.get('acconto_imposta_perc', 1.0)}"></div>
            <div class="field"><label>Acconto — 1ª rata, quota a giugno{_info("0,40 = 40 % con il saldo di giugno, il resto al 30/11. A 0 va tutto sulla scadenza di novembre.")}</label>
              <input type="number" step="0.01" id="f_acc_rata"
                     value="{p.get('acconto_prima_rata_perc', 0.40)}"></div>
            <div class="field"><label>Limite fatturato annuo (€)</label>
              <input type="number" step="0.01" id="f_limite" value="{p['limite_fatturato_anno']}"></div>
          </div>
          <div class="field-group">
            <div class="field"><label>Bollo — soglia (€)</label>
              <input type="number" step="0.01" id="f_bollo_soglia" value="{p['bollo_soglia']}"></div>
            <div class="field"><label>Bollo — importo (€)</label>
              <input type="number" step="0.01" id="f_bollo_importo" value="{p['bollo_importo']}"></div>
          </div>
        </div>

        <div class="card">
          <div class="card-head"><div class="eyebrow">Tariffa</div></div>
          <div class="field"><label>Tariffa giornaliera (€){_info("Quanto vale una giornata da 8 ore. È il prezzo che il timesheet propone quando precompila la fattura di fine mese: giornate × tariffa, una riga sola.")}</label>
            <input type="number" step="0.01" min="0" id="f_tariffa"
                   value="{p.get('tariffa_giornaliera', 250)}"></div>
        </div>

        <div class="card">
          <div class="card-head"><div class="eyebrow">Accantonamento</div></div>
          <p class="small muted mb-3">
            Questi tre valori determinano lo scarto tra il dovuto matematico e
            quanto l'app ti consiglia di mettere da parte.
          </p>
          <div class="field-group">
            <div class="field"><label>Margine di sicurezza{_info("0,10 = 10 % in più del dovuto.")}</label>
              <input type="number" step="0.01" id="f_margine" value="{p.get('margine_sicurezza', 0.10)}"></div>
            <div class="field"><label>Scenario preferito</label>
              <select id="f_scenario">{scen_opts}</select>
              <div class="hint">Quello mostrato per primo.</div></div>
          </div>
          <div class="field-group">
            <div class="field"><label>Costi fissi annui (€){_info("Commercialista, PEC, bolli, commissioni.")}</label>
              <input type="number" step="0.01" id="f_costi" value="{p.get('costi_fissi_annui', 0)}"></div>
            <div class="field"><label>Fatturato atteso annuo (€){_info("Su cui spalmare i costi fissi. A zero li stima dall'incassato dell'anno.")}</label>
              <input type="number" step="0.01" id="f_atteso" value="{p.get('fatturato_atteso_anno', 0)}"></div>
          </div>
        </div>

        <div class="actions">
          <button type="button" class="btn" onclick="onSalva()">Salva parametri</button>
        </div>
      </div>

      <div class="stack">
        <div class="card">
          <div class="card-head"><div class="eyebrow">Effetto su 1.000 € incassati</div></div>
          <div class="rows detail">
            {righe_anteprima}
          </div>
          <div class="small muted mt-3">
            Salva per aggiornare l'anteprima.
          </div>
        </div>
      </div>
    </div>

    <div id="toast" class="toast"></div>
    <script>
    function toast(msg, cls) {{
      const t = document.getElementById('toast');
      t.textContent = msg; t.className = 'toast show ' + (cls || '');
      setTimeout(()=>{{t.className='toast '+(cls||'')}}, 2200);
    }}
    async function onSalva() {{
      const g = id => Number(document.getElementById(id).value);
      const v = id => document.getElementById(id).value;
      const body = {{
        regime: v('f_regime'),
        coeff_ateco: g('f_coeff'),
        aliquota_imposta: g('f_aliq_imp'),
        aliquota_inps: g('f_aliq_inps'),
        aliquota_acconto: g('f_aliq_acc'),
        acconto_imposta_perc: g('f_acc_imp'),
        acconto_prima_rata_perc: g('f_acc_rata'),
        bollo_soglia: g('f_bollo_soglia'),
        bollo_importo: g('f_bollo_importo'),
        limite_fatturato_anno: g('f_limite'),
        data_apertura_piva: v('f_data_apertura'),
        anno_fine_regime_agevolato: g('f_anno_fine'),
        margine_sicurezza: g('f_margine'),
        costi_fissi_annui: g('f_costi'),
        fatturato_atteso_anno: g('f_atteso'),
        scenario_preferito: v('f_scenario'),
        tariffa_giornaliera: g('f_tariffa'),
      }};
      try {{
        const r = await fetch('/fatture/api/parametri', {{
          method: 'PATCH', headers: {{'Content-Type': 'application/json'}},
          body: JSON.stringify(body),
        }});
        const j = await r.json();
        if (!r.ok) {{ toast(j.error || 'Errore', 'err'); return; }}
        toast('Parametri aggiornati', 'ok');
        setTimeout(() => location.reload(), 700);
      }} catch (e) {{ toast('Errore rete: '+e.message, 'err'); }}
    }}
    </script>
    '''
    return _render(body, eyebrow="Parametri fiscali",
                   title_html='Parametri <em>fiscali</em>', breadcrumb=breadcrumb)


@fatture_bp.get("/api/parametri")
def api_parametri_get():
    sb, err = _supabase_or_error()
    if err: return jsonify({"error": "supabase not configured"}), 503
    return jsonify(_get_parametri(sb))


@fatture_bp.patch("/api/parametri")
def api_parametri_update():
    sb, err = _supabase_or_error()
    if err: return jsonify({"error": "supabase not configured"}), 503
    data = request.get_json(silent=True) or {}
    payload = {k: data[k] for k in PARAMETRI_CAMPI if k in data}
    errore_validazione = _valida_parametri(payload)
    if errore_validazione:
        return jsonify({"error": errore_validazione}), 400
    try:
        r = sb.table("b2f_parametri_fiscali").update(payload).eq("id", 1).execute()
        return jsonify(r.data[0] if r.data else {"id": 1})
    except Exception as e:
        msg = str(e)
        # I parametri di accantonamento vivono su colonne aggiunte da una
        # migrazione. Se non e' stata ancora eseguita, dirlo invece di
        # rilanciare l'errore grezzo del database.
        if "column" in msg.lower() and any(c in msg for c in acc.PARAMETRI_CAMPI):
            return jsonify({
                "error": "Colonne mancanti su b2f_parametri_fiscali: esegui le "
                         "migrazioni documentate nel README nell'SQL Editor di "
                         "Supabase."
            }), 409
        return jsonify({"error": msg[:200]}), 500


# ---------------------------------------------------------------------------
# CRUD spese P.IVA
# ---------------------------------------------------------------------------

def _card_rivalsa(rivalsa: float, conto: dict) -> str:
    """
    Quanta parte del saldo P.IVA e' rivalsa INPS incassata dai clienti.

    Dentro il saldo la rivalsa e' invisibile: e' arrivata insieme al
    corrispettivo, in un unico bonifico, e da li' in poi e' un numero
    solo. Ma non e' un ricavo tuo: e' la parte destinata al contributo
    previdenziale, e deve restare su questo conto finche' non la versi.
    Qui si dice quanta e', e che e' gia' compresa nell'accantonamento —
    non e' una quota da mettere da parte in piu'.
    """
    saldo = float(conto.get("saldo") or 0)
    # Niente percentuale sul saldo: accanto a "rivalsa 4 %" un secondo
    # numero in percentuale si legge come l'aliquota e confonde. Meglio
    # il confronto in euro, che e' quello che serve davvero.
    quota = (f' Sul conto ci sono in tutto € {_fmt_eur(saldo)}.'
             if saldo > 0 else "")
    return f'''
    <div class="card mb-3">
      <div class="card-head">
        <div class="eyebrow">Rivalsa INPS incassata</div>
        <span class="chip accent">€ {_fmt_eur(rivalsa)}</span>
      </div>
      <p class="small muted">
        È la quota di rivalsa INPS 4 % già entrata su questo conto insieme
        agli incassi delle fatture.{quota} Non è un ricavo: è la parte del
        corrispettivo destinata al contributo previdenziale, e va lasciata
        qui fino al versamento.
      </p>
      <details class="explain mt-3">
        <summary>Devo accantonarla a parte?</summary>
        <p class="small muted mt-2">
          No: è già dentro. Nel forfettario l'imponibile è il corrispettivo
          intero, rivalsa compresa, quindi l'INPS calcolata
          sull'accantonamento (circa il 17,5 % del lordo) è più alta della
          rivalsa stessa (3,85 % del lordo) e la contiene. Metterla da parte
          una seconda volta significherebbe accantonare due volte lo stesso
          denaro. La ripartizione dell'incasso non scende mai sotto la
          rivalsa proprio per questo.
        </p>
      </details>
    </div>'''


def _movimento_label(m: dict) -> str:
    tipo = m.get("tipo") or "uscita"
    segno = {"entrata": "+", "uscita": "−", "giroconto": "⇄"}.get(tipo, "")
    cls = {"entrata": "pos", "uscita": "neg", "giroconto": ""}.get(tipo, "")
    return segno, cls


@fatture_bp.get("/spese-piva")
def spese_piva_list():
    sb, err = _supabase_or_error()
    breadcrumb = [("Fatture", "/fatture"), ("Situazione fiscale", "/fatture/situazione"),
                  ("Spese P.IVA", "")]
    if err:
        return _render(err, breadcrumb=breadcrumb, fab=("Nuovo movimento", "/fatture/spese-piva/nuova"))

    anno_default = date.today().year
    anno = request.args.get("anno", type=int) or anno_default
    categoria = request.args.get("categoria") or ""
    tipo = request.args.get("tipo") or ""

    try:
        q = (sb.table("b2f_spese_piva").select("*")
               .gte("data", f"{anno}-01-01").lte("data", f"{anno}-12-31")
               .order("data", desc=True))
        if categoria:
            q = q.eq("categoria", categoria)
        if tipo:
            q = q.eq("tipo", tipo)
        rows = (q.execute()).data or []
    except Exception as e:
        return _render(f'<div class="notice err">Errore: {str(e)[:200]}</div>', breadcrumb=breadcrumb)

    cat_opts = "".join(
        f'<option value="{k}"{" selected" if categoria==k else ""}>{lbl}</option>'
        for k, lbl in CATEGORIE_SPESE_PIVA
    )
    tipo_opts = "".join(
        f'<option value="{k}"{" selected" if tipo==k else ""}>{lbl}</option>'
        for k, lbl in TIPI_SPESE_PIVA
    )

    anno_o = "".join(f'<option value="{y}"{" selected" if y == anno else ""}>{y}</option>'
                     for y in range(anno_default, anno_default - 6, -1))
    toolbar = f'''
    <div class="toolbar">
      <select class="select-pill" aria-label="Anno"
        onchange="const u=new URL(location.href);u.searchParams.set('anno',this.value);location.href=u">
        {anno_o}
      </select>
      <select class="select-pill" aria-label="Categoria"
        onchange="const u=new URL(location.href);if(this.value){{u.searchParams.set('categoria',this.value)}}else{{u.searchParams.delete('categoria')}};location.href=u">
        <option value="">Tutte le categorie</option>{cat_opts}
      </select>
      <select class="select-pill" aria-label="Tipo"
        onchange="const u=new URL(location.href);if(this.value){{u.searchParams.set('tipo',this.value)}}else{{u.searchParams.delete('tipo')}};location.href=u">
        <option value="">Tutti i tipi</option>{tipo_opts}
      </select>
    </div>
    '''

    entrate = sum(float(r.get("importo") or 0) for r in rows if r.get("tipo") == "entrata")
    uscite = sum(float(r.get("importo") or 0) for r in rows if r.get("tipo") == "uscita")
    # I giroconti escono dal conto P.IVA verso il personale: non sono una
    # spesa (restano tuoi), ma il saldo deve vederli uscire, altrimenti
    # mostrerebbe soldi che sull'altro conto ci sono gia'.
    girati = sum(float(r.get("importo") or 0) for r in rows if r.get("tipo") == "giroconto")
    tot = entrate - uscite - girati

    # Il saldo qui sopra e' quello dell'anno filtrato. Il conto pero' non
    # riparte da zero a gennaio: dal secondo anno in poi i due numeri
    # divergono, e il primo non e' quello che c'e' in banca.
    conto = saldo_piva(sb)
    rivalsa = float(conto.get("rivalsa_incassata") or 0)
    riga_conto = ""
    if conto.get("disponibile"):
        rivalsa_hint = (f'di cui € {_fmt_eur(rivalsa, 0)} di rivalsa INPS incassata'
                        if rivalsa > 0 else "tutti i movimenti fino a oggi")
        riga_conto = f'''
      <div class="card"><div class="stat">
        <div class="val tnum {"pos" if conto["saldo"] >= 0 else "neg"}">€ {_fmt_eur(conto["saldo"], 0)}</div>
        <div class="lbl">Saldo del conto, oggi</div>
        <div class="hint">{rivalsa_hint}</div></div></div>'''

    riepilogo = f'''
    <div class="grid kpi lead mb-3">
      {riga_conto}
      <div class="card"><div class="stat">
        <div class="val tnum {"pos" if tot >= 0 else "neg"}">€ {_fmt_eur(tot, 0)}</div>
        <div class="lbl">Movimento netto {anno}</div>
        <div class="hint">al netto dei giroconti</div></div></div>
      <div class="card"><div class="stat sm">
        <div class="val tnum pos">€ {_fmt_eur(entrate, 0)}</div>
        <div class="lbl">Entrate</div></div></div>
      <div class="card"><div class="stat sm">
        <div class="val tnum neg">€ {_fmt_eur(uscite, 0)}</div>
        <div class="lbl">Uscite</div></div></div>
      <div class="card"><div class="stat sm">
        <div class="val tnum">€ {_fmt_eur(girati, 0)}</div>
        <div class="lbl">Girati al personale</div></div></div>
    </div>
    {_card_rivalsa(rivalsa, conto) if rivalsa > 0 else ""}
    '''

    if not rows:
        body = f'''{riepilogo}{toolbar}
        <div class="empty">
          {_icon("wallet")}
          <div class="t">Nessun movimento per il {anno}</div>
          <div class="s">Aggiungi il primo con il pulsante in basso.</div>
        </div>'''
    else:
        cat_lbl = dict(CATEGORIE_SPESE_PIVA)
        items = []
        for m in rows:
            segno, cls = _movimento_label(m)
            cat = _esc(cat_lbl.get(m.get("categoria"), m.get("categoria") or "—"))
            items.append(f'''
            <a class="item" href="/fatture/spese-piva/{m["id"]}">
              <span class="ico {cls or "neutral"}">{_icon("wallet")}</span>
              <span class="body">
                <span class="n">{_esc(m.get("descrizione") or "—")}</span>
                <span class="m">{_fmt_date(m.get("data"))} · {cat}</span>
              </span>
              <span class="end">
                <span class="amt tnum {cls}">{segno} € {_fmt_eur(m.get("importo"))}</span>
              </span>
            </a>''')
        body = f'{riepilogo}{toolbar}<div class="list">{"".join(items)}</div>'

    return _render(body, eyebrow="Movimenti P.IVA", title_html='Movimenti <em>P.IVA</em>',
                   breadcrumb=breadcrumb, fab=("Nuovo movimento", "/fatture/spese-piva/nuova"))


def _movimento_form_html(m: dict | None = None, collegamento: dict | None = None) -> str:
    m = m or {}
    v = lambda k, d="": (m.get(k) if m.get(k) is not None else d)
    tipo_current = m.get("tipo") or "uscita"
    tipo_opts = "".join(
        f'<option value="{k}"{" selected" if k==tipo_current else ""}>{lbl}</option>'
        for k, lbl in TIPI_SPESE_PIVA
    )
    cat_current = m.get("categoria") or ""
    cat_opts = "".join(
        f'<option value="{k}"{" selected" if k==cat_current else ""}>{lbl}</option>'
        for k, lbl in CATEGORIE_SPESE_PIVA
    )
    mid = m.get("id") or ""
    is_edit = bool(mid)
    submit_lbl = "Aggiorna" if is_edit else "Registra movimento"

    # Un movimento che fa parte di un giroconto non si tocca da qui: si
    # può solo, quando questa e' l'origine dello spostamento, eliminarlo
    # (cosa che porta via anche la contropartita). Le guardie vere stanno
    # nell'endpoint; qui e' solo per non lasciare che l'utente provi un
    # salvataggio che verrebbe comunque rifiutato.
    bloccato = bool(collegamento)
    delete_ok = (not bloccato) or bool(collegamento.get("delete_ok"))
    ro = " readonly disabled" if bloccato else ""

    avviso = (f'<div class="notice info mb-4">{collegamento["msg"]}</div>'
              if collegamento else "")
    submit_btn = ("" if bloccato else
                  f'<button type="button" class="btn" onclick="onSubmit({mid or "null"})">{submit_lbl}</button>')
    delete_btn = (f'<button type="button" class="btn danger" onclick="onElimina({mid})">Elimina</button>'
                  if is_edit and delete_ok else "")
    torna_btn = ('<a class="btn ghost" href="/fatture/spese-piva">Torna ai movimenti</a>'
                 if bloccato and not delete_ok else "")

    return f'''
    <div class="narrow">
    {avviso}
    <div class="card">
      <div class="field-group">
        <div class="field"><label>Data</label>
          <input type="date" id="f_data" value="{v('data', date.today().isoformat())}"{ro}></div>
        <div class="field"><label>Importo (€)</label>
          <input type="number" step="0.01" inputmode="decimal" id="f_importo" value="{v('importo', 0)}"{ro}></div>
      </div>
      <div class="field"><label>Tipo</label>
        <select id="f_tipo"{ro}>{tipo_opts}</select></div>
      <div class="field"><label>Descrizione</label>
        <input id="f_descrizione" value="{_esc(v('descrizione'))}"{ro}></div>
      <div class="field-group">
        <div class="field"><label>Categoria</label>
          <select id="f_categoria"{ro}><option value="">—</option>{cat_opts}</select></div>
        <div class="field"><label>Sottocategoria</label>
          <input id="f_sottocategoria" value="{_esc(v('sottocategoria'))}"{ro}></div>
      </div>
      <div class="field inline">
        <input type="checkbox" id="f_ricorrente" {"checked" if m.get("ricorrente") else ""}{ro}>
        <label for="f_ricorrente">Movimento ricorrente</label>
      </div>
      <div class="field"><label>Note</label>
        <textarea id="f_note"{ro}>{_esc(m.get('note'))}</textarea></div>
      <div class="actions">
        {submit_btn}
        {delete_btn}
        {torna_btn}
      </div>
    </div>
    </div>
    <div id="toast" class="toast"></div>
    <script>
    function toast(msg, cls) {{
      const t = document.getElementById('toast');
      t.textContent = msg; t.className = 'toast show ' + (cls || '');
      setTimeout(()=>{{t.className='toast '+(cls||'')}}, 2200);
    }}
    function readForm() {{
      return {{
        data: document.getElementById('f_data').value,
        tipo: document.getElementById('f_tipo').value,
        importo: Number(document.getElementById('f_importo').value || 0),
        descrizione: document.getElementById('f_descrizione').value.trim(),
        categoria: document.getElementById('f_categoria').value || null,
        sottocategoria: document.getElementById('f_sottocategoria').value.trim() || null,
        ricorrente: document.getElementById('f_ricorrente').checked,
        note: document.getElementById('f_note').value.trim() || null,
      }};
    }}
    async function onSubmit(mid) {{
      const body = readForm();
      const isNew = !mid;
      const url = isNew ? '/fatture/api/spese-piva' : '/fatture/api/spese-piva/'+mid;
      const method = isNew ? 'POST' : 'PATCH';
      try {{
        const r = await fetch(url, {{
          method, headers: {{'Content-Type':'application/json'}}, body: JSON.stringify(body),
        }});
        const j = await r.json();
        if (!r.ok) {{ toast(j.error || 'Errore', 'err'); return; }}
        toast(isNew ? 'Movimento registrato' : 'Aggiornato', 'ok');
        setTimeout(()=>{{ location.href = '/fatture/spese-piva'; }}, 500);
      }} catch (e) {{ toast('Errore: '+e.message, 'err'); }}
    }}
    async function onElimina(mid) {{
      if (!confirm('Eliminare questo movimento?')) return;
      try {{
        const r = await fetch('/fatture/api/spese-piva/'+mid, {{method:'DELETE'}});
        if (!r.ok) {{ toast('Errore', 'err'); return; }}
        toast('Eliminato', 'ok');
        setTimeout(()=>{{ location.href = '/fatture/spese-piva'; }}, 500);
      }} catch (e) {{ toast('Errore: '+e.message, 'err'); }}
    }}
    </script>
    '''


@fatture_bp.get("/spese-piva/nuova")
def spesa_piva_new():
    breadcrumb = [("Fatture", "/fatture"), ("Situazione fiscale", "/fatture/situazione"),
                  ("Spese P.IVA", "/fatture/spese-piva"), ("Nuovo", "")]
    return _render(_movimento_form_html(None), eyebrow="Nuovo movimento",
                   title_html='<em>Nuovo</em> movimento', breadcrumb=breadcrumb)


@fatture_bp.get("/spese-piva/<int:mid>")
def spesa_piva_edit(mid):
    sb, err = _supabase_or_error()
    breadcrumb = [("Fatture", "/fatture"), ("Situazione fiscale", "/fatture/situazione"),
                  ("Spese P.IVA", "/fatture/spese-piva"), (str(mid), "")]
    if err:
        return _render(err, breadcrumb=breadcrumb)
    try:
        r = sb.table("b2f_spese_piva").select("*").eq("id", mid).single().execute()
        m = r.data
    except Exception as e:
        return _render(f'<div class="notice err">{str(e)[:200]}</div>', breadcrumb=breadcrumb)

    origine = _origine_ripartizione(sb, mid)
    if origine:
        collegamento = {
            "msg": (f'Questo movimento fa parte della ripartizione della fattura '
                    f'{origine.get("numero") or origine.get("id")}. Per toccarlo '
                    f'annulla la ripartizione dalla fattura: così spariscono '
                    f'entrambe le righe.'),
            "delete_ok": False,
        }
    elif m.get("giroconto_personale_id"):
        collegamento = {
            "msg": ("Questo movimento ha generato un'entrata sul conto "
                    "personale. Tipo e importo non si possono cambiare da "
                    "qui: eliminalo per annullare entrambe le righe."),
            "delete_ok": True,
        }
    else:
        collegamento = None

    return _render(_movimento_form_html(m, collegamento), eyebrow="Movimento",
                   title_html=f'<em>{_esc((m.get("descrizione") or "Movimento")[:20])}</em>',
                   breadcrumb=breadcrumb)


@fatture_bp.get("/api/spese-piva")
def api_spese_piva_list():
    sb, err = _supabase_or_error()
    if err: return jsonify({"error": "supabase not configured"}), 503
    anno = request.args.get("anno", type=int)
    categoria = request.args.get("categoria")
    tipo = request.args.get("tipo")
    try:
        q = sb.table("b2f_spese_piva").select("*").order("data", desc=True)
        if anno:
            q = q.gte("data", f"{anno}-01-01").lte("data", f"{anno}-12-31")
        if categoria:
            q = q.eq("categoria", categoria)
        if tipo:
            q = q.eq("tipo", tipo)
        r = q.execute()
        return jsonify(r.data or [])
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


def _movimento_payload(data: dict) -> dict:
    out = {}
    for k in MOVIMENTO_CAMPI:
        if k in data:
            v = data[k]
            if isinstance(v, str):
                v = v.strip() or None
            out[k] = v
    return out


def _origine_ripartizione(sb, mid: int) -> dict | None:
    """
    Se questo movimento e' la meta' P.IVA della ripartizione automatica
    di un incasso (fatture/giroconto.py), la fattura a cui appartiene.

    Il legame piu' debole di `spesa_piva_id` (la semplice registrazione
    dell'incasso) di per se' si scioglie apposta eliminando il movimento
    da qui — vedi `_stacca_da_fattura`. Ma se sopra quella registrazione
    e' *gia'* stato eseguito un giroconto (`giroconto_piva_id` valorizzato
    sulla stessa fattura), cancellare la riga "fatturato" lascerebbe la
    meta' di uscita del giroconto nel libro P.IVA senza piu' un incasso
    a giustificarla — quindi da quel momento in poi va protetta anche lei.
    """
    try:
        r = (sb.table("b2f_fatture").select("id, numero")
               .eq("giroconto_piva_id", mid).limit(1).execute())
        righe = r.data or []
        if righe:
            return righe[0]
        r2 = (sb.table("b2f_fatture").select("id, numero, giroconto_piva_id")
                .eq("spesa_piva_id", mid).limit(1).execute())
        righe2 = r2.data or []
        if righe2 and righe2[0].get("giroconto_piva_id"):
            return righe2[0]
        return None
    except Exception:
        return None


def _stacca_da_fattura(sb, mid: int) -> None:
    """
    Se il movimento appena eliminato era la registrazione dell'incasso di
    una fattura, scioglie il collegamento: altrimenti la fattura resta a
    puntare a una riga sparita, "già registrata" per sempre.
    """
    try:
        sb.table("b2f_fatture").update({"spesa_piva_id": None}) \
          .eq("spesa_piva_id", mid).execute()
    except Exception:
        pass


@fatture_bp.post("/api/spese-piva")
def api_spesa_piva_create():
    sb, err = _supabase_or_error()
    if err: return jsonify({"error": "supabase not configured"}), 503
    data = request.get_json(silent=True) or {}
    for k in ("data", "importo", "tipo", "descrizione"):
        if not data.get(k) and data.get(k) != 0:
            return jsonify({"error": f"campo mancante: {k}"}), 400
    if data.get("tipo") not in ("entrata", "uscita", "giroconto"):
        return jsonify({"error": "tipo non valido"}), 400
    payload = _movimento_payload(data)
    try:
        r = sb.table("b2f_spese_piva").insert(payload).execute()
        riga = r.data[0] if r.data else {}
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500

    id_piva = riga.get("id")
    if payload.get("tipo") != "giroconto" or not id_piva:
        return jsonify(riga)

    # Un giroconto sposta soldi anche sul conto personale: senza questa
    # riga il denaro sparirebbe dal conto P.IVA senza arrivare da nessuna
    # parte. Stessa garanzia della ripartizione automatica di una fattura
    # (fatture/giroconto.py), solo innescata a mano invece che dall'incasso.
    from spese import dati as personale
    esito = personale.crea(sb, {
        "data":              payload.get("data"),
        "tipo":              "entrata",
        "importo":           payload.get("importo"),
        "descrizione":       f'Giroconto da P.IVA — {payload.get("descrizione")}',
        "metodo_pagamento":  "Giroconto",
        "categoria_link_id": personale.link_categoria(
            sb, personale.CATEGORIA_GIROCONTO),
    })
    if esito.get("error") or not esito.get("id"):
        try:
            sb.table("b2f_spese_piva").delete().eq("id", id_piva).execute()
        except Exception:
            pass
        msg = esito.get("error") or "nessun id di ritorno"
        return jsonify({"error": f"errore sul conto personale: {msg}"}), 500

    try:
        upd = (sb.table("b2f_spese_piva")
                 .update({"giroconto_personale_id": esito["id"]})
                 .eq("id", id_piva).execute())
        riga = upd.data[0] if upd.data else riga
    except Exception as e:
        # Senza il collegamento le due righe si perdono di vista: meglio
        # annullare tutto che lasciarle scollegate.
        try:
            sb.table("b2f_spese_piva").delete().eq("id", id_piva).execute()
        except Exception:
            pass
        try:
            personale.elimina(sb, esito["id"])
        except Exception:
            pass
        msg = str(e)
        if "column" in msg.lower() and "giroconto_personale_id" in msg:
            return jsonify({"error": (
                "Manca la colonna giroconto_personale_id su b2f_spese_piva: "
                "esegui la migrazione documentata nel README.")}), 409
        return jsonify({"error": f"collegamento fallito: {msg[:200]}"}), 500

    return jsonify(riga)


@fatture_bp.patch("/api/spese-piva/<int:mid>")
def api_spesa_piva_update(mid):
    sb, err = _supabase_or_error()
    if err: return jsonify({"error": "supabase not configured"}), 503

    origine = _origine_ripartizione(sb, mid)
    if origine:
        return jsonify({"error": (
            f'Questo movimento fa parte della ripartizione della fattura '
            f'{origine.get("numero") or origine.get("id")}: per cambiarlo '
            f'annulla la ripartizione dalla fattura, così spariscono entrambe '
            f'le righe.')}), 409

    try:
        r = sb.table("b2f_spese_piva").select("giroconto_personale_id") \
              .eq("id", mid).single().execute()
        collegato_manuale = (r.data or {}).get("giroconto_personale_id")
    except Exception:
        collegato_manuale = None

    data = request.get_json(silent=True) or {}
    if collegato_manuale and ("tipo" in data or "importo" in data):
        return jsonify({"error": (
            "Questo movimento ha generato un'entrata sul conto personale: "
            "tipo e importo non si possono cambiare da qui, per non "
            "disallineare i due conti. Eliminalo e registrane uno nuovo se "
            "serve un altro importo.")}), 409

    payload = _movimento_payload(data)
    try:
        r = sb.table("b2f_spese_piva").update(payload).eq("id", mid).execute()
        return jsonify(r.data[0] if r.data else {"id": mid})
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500


@fatture_bp.delete("/api/spese-piva/<int:mid>")
def api_spesa_piva_delete(mid):
    sb, err = _supabase_or_error()
    if err: return jsonify({"error": "supabase not configured"}), 503

    origine = _origine_ripartizione(sb, mid)
    if origine:
        return jsonify({"error": (
            f'Questo movimento fa parte della ripartizione della fattura '
            f'{origine.get("numero") or origine.get("id")}: annulla la '
            f'ripartizione dalla fattura, così spariscono entrambe le righe.'
        )}), 409

    try:
        r = sb.table("b2f_spese_piva").select("giroconto_personale_id") \
              .eq("id", mid).single().execute()
        id_personale = (r.data or {}).get("giroconto_personale_id")
    except Exception:
        id_personale = None

    if id_personale:
        # Il giroconto e' nato da qui: eliminarlo toglie anche la
        # contropartita sul personale, altrimenti resterebbe un'entrata
        # senza piu' nessuna uscita dal conto P.IVA a spiegarla.
        from spese import dati as personale
        esito = personale.elimina(sb, id_personale)
        if esito.get("error"):
            return jsonify({"error": (
                "Non sono riuscito a togliere l'entrata collegata sul conto "
                f"personale: {esito['error']}")}), 500

    try:
        sb.table("b2f_spese_piva").delete().eq("id", mid).execute()
    except Exception as e:
        return jsonify({"error": str(e)[:200]}), 500

    _stacca_da_fattura(sb, mid)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _render(content: str, eyebrow: str = "Situazione fiscale",
            title_html: str = 'Situazione <em>fiscale</em>',
            breadcrumb=None, fab=None, actions_html: str = "") -> Response:
    html = render_page(
        section="fatture", eyebrow=eyebrow, title_html=title_html,
        content=content, breadcrumb=breadcrumb, fab=fab,
        actions_html=actions_html,
    )
    return Response(html, mimetype="text/html")
